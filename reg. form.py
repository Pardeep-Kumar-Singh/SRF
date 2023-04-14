from tkinter import *
from tkcalendar import Calendar
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl , xlrd
from openpyxl import Workbook
import pathlib

background="#06283D"
framebg="#EDEDED"
framefg="#06283D"

root = Tk() #create object 
root.title("Student Registration System") #give title
root.geometry("1250x700+210+100") #give size of a output screen
root.configure(bg=background)

#create a excel file and save data
file=pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Roll No."
    sheet['B1']="Name"
    sheet['C1']="Class"
    sheet['D1']="Gender"
    sheet['E1']="DOB"
    sheet['F1']="Date of Registration"
    sheet['G1']="E-mail"
    sheet['H1']="Skill"
    sheet['I1']="Father Name"
    sheet['J1']="Mother Name"
    sheet['K1']="Father's Occupation"
    sheet['L1']="Mother's Occupation"

    file.save('Student_data.xlsx')

#Exit window
def Exit():
    root.destroy()


#showimage
def showimage():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),title="Select image file",filetype=(("JPG FILE","*.jpg"),("PNG FILE","*.png"),("ALL FILES","*.txt")))
    img=(Image.open(filename))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

#Roll NO.
def roll_no():
    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=1).value

    try:
        rollno.set(max_row_value+1)

    except:
        Rollno.set("200130800000")


#Clear
def Clear():
    global img
    Name.set('')
    DOB.set('')
    mail.set('')
    Skills.set('')
    Fname.set('')
    Mname.set('')
    F_occ_name.set('')
    M_occ_name.set('')
    Semester.set("Select Semester")

    roll_no()

    saveButton.config(state='normal')

    img1=PhotoImage(file="upload.png")
    #lbl=config(image=img1)
    lbl.image=img1

    img=""

    
#Save
def Save():
    R1=Rollno.get()
    N1=Name.get()
    S1=Semester.get()
    try:
        G1=gender
    except:
        messagebox.showerror("error","Select Gender")
    D2=DOB.get()
    D1=Date.get()
    email=mail.get()
    S2=Skills.get()
    fathername=Fname.get()
    mothername=Mname.get()
    F1=F_occ_name.get()
    M1=M_occ_name.get()

    if N1=="" or S1=="Select Class" or D2=="" or email=="" or S2=="" or fathername=="" or mothername=="" or F1=="" or M1=="" :
        messagebox.showerror("error","Few Data is missing!")
    else:
        file=openpyxl.load_workbook("Student_data.xlsx")
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=N1)
        sheet.cell(column=3,row=sheet.max_row,value=S1)
        sheet.cell(column=4,row=sheet.max_row,value=G1)
        sheet.cell(column=5,row=sheet.max_row,value=D2)
        sheet.cell(column=6,row=sheet.max_row,value=D1)
        sheet.cell(column=7,row=sheet.max_row,value=email)
        sheet.cell(column=8,row=sheet.max_row,value=S2)
        sheet.cell(column=9,row=sheet.max_row,value=fathername)
        sheet.cell(column=10,row=sheet.max_row,value=mothername)
        sheet.cell(column=11,row=sheet.max_row,value=F1)
        sheet.cell(column=12,row=sheet.max_row,value=M1)

        file.save("Student_data.xlsx")

        try:
            img.save("Student Images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info","Profilr picture is not available!!!")
        messagebox.showinfo("info","Succesfully data entered!!!")

        Clear()
        
###search####
def search():
    text = Search.get()
    Clear()
    #saveButton.config(state="disable")

    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value==int(text):
            name=row[0]
            #print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid","Invalid registration number!!!")
            
    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=7).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value

####    print(x1)
####    print(x2)
####    print(x3)
####    print(x4)
####    print(x5)
####    print(x6)
####    print(x7)
####    print(x8)
####    print(x9)
####    print(x10)
####    print(x11)
####    print(x12)

    Rollno.set(x1)
    Name.set(x2)
    Semester.set(x3)

    if x4=='female':
        R2.select()
    else:
        R1.select()

    DOB.set(x5)
    Date.set(x6)
    mail.set(x7)
    Skills.set(x8)
    Fname.set(x9)
    Mname.set(x10)
    F_occ_name.set(x11)
    M_occ_name.set(x12)

    img=(Image.open("Student Images/"+str(x1)+".jpg"))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

    
##Update

    
def Update():
    R1=Rollno.get()
    N1=Name.get()
    S1=Semester.get()
    selection()
    G1=gender
    D2=DOB.get()
    D1=Date.get()
    email=mail.get()
    S2=Skills.get()
    fathername=Fname.get()
    mothername=Mname.get()
    F1=F_occ_name.get()
    M1=M_occ_name.get()
    
    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value==R1:
            name=row[0]
            print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

            #print(reg_number)

    #sheet.cell(column=1,row=int(reg_number),value=R1)
    sheet.cell(column=2,row=int(reg_number),value=N1)
    sheet.cell(column=3,row=int(reg_number),value=S1)
    sheet.cell(column=4,row=int(reg_number),value=G1)
    sheet.cell(column=5,row=int(reg_number),value=D2)
    sheet.cell(column=6,row=int(reg_number),value=D1)
    sheet.cell(column=7,row=int(reg_number),value=email)
    sheet.cell(column=8,row=int(reg_number),value=S2)
    sheet.cell(column=9,row=int(reg_number),value=fathername)
    sheet.cell(column=10,row=int(reg_number),value=mothername)
    sheet.cell(column=11,row=int(reg_number),value=F1)
    sheet.cell(column=12,row=int(reg_number),value=M1)

    file.save("Student_data.xlsx")

    try:
        img.save("Student Images/"+str(R1)+".jpg")
    except:
        pass
        
    messagebox.showinfo("Update","Update Succesfully")

    Clear()

#gender
def selection():
    global gender
    value=radio.get()
    if value==1:
        gender="Male"
    else:
        gender="female"

#DOB pick
def pick_date(event):
    global cal,date_window

    date_window=Toplevel()
    date_window.grab_set()
    date_window.title("Choose Date of Birth")
    date_window.geometry('250x220+590+370')
    cal= calendar(date_window,selectmode="day",date_pattern="mm/dd/yy")
    cal.place(x=0, y=0)

    submit_btn = Button(date_window,text="Submit", command=grab_date)
    submit_btn.place(x=80, y=190)

def grab_date():
    dob_entry.delete(0,END)
    dob_entry.insert(0, cal.get_date())
    date_window.destroy()

#top frame
Label(root,text="Email: studentregistration@gmail.com",width=10,height=3,bg="#f0687c",font='arial 10 bold',anchor='e').pack(side=TOP,fill=X)
Label(root,text="STUDENT REGISTRATION",width=10,height=2,bg="#c36464",fg="#fff",font='arial 20 bold').pack(side=TOP,fill=X)

#search box to update
Search=StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font='arial 20').place(x=820,y=70)
imageicon3=PhotoImage(file="search.png")#search icon image
Srch=Button(root,text="Search",compound=LEFT,image=imageicon3,width=123,bg="#68ddfa",font='arial 13 bold',command=search)
Srch.place(x=1060,y=66)

imageicon4=PhotoImage(file="layer5.png")
update_button=Button(root,image=imageicon4,bg='#c36464',command=Update)
update_button.place(x=110,y=64)

#Roll_no and Date
Label(root,text="Roll No:",font="arial 13",fg=framebg,bg=background).place(x=30,y=150)
Label(root,text="Date:",font="arial 13",fg=framebg,bg=background).place(x=500,y=150)

Rollno=IntVar()
Date=StringVar()

rollno_entry=Entry(root,textvariable=Rollno,width=15,font='arial 10')
rollno_entry.place(x=160,y=150)

roll_no()

today = date.today()
d1=today.strftime("%d/%m/%y")
date_entry=Entry(root,textvariable=Date,width=15,font='arial 10')
date_entry.place(x=550,y=150)

Date.set(d1)

#student details
obj=LabelFrame(root,text="Student's Detail",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Full Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Date of Birth:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)

Label(obj,text="Semester:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj,text="E-mail:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj,text="Skills:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)

Name=StringVar()
name_entry=Entry(obj,textvariable=Name,width=20,font="arial 10")
name_entry.place(x=160,y=50)

DOB=StringVar()
dob_entry=Entry(obj,textvariable=DOB,width=20,font="arial 10")
dob_entry.place(x=160,y=100)
dob_entry.insert(0,"dd/mm/yy")
dob_entry.bind("<1>",pick_date)

radio=IntVar()
R1=Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=150,y=150)

R2=Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
R2.place(x=200,y=150)

mail=StringVar()
email=Entry(obj,textvariable=mail,width=20,font="arial 10")
email.place(x=630,y=100)

Skills=StringVar()
skills_entry=Entry(obj,textvariable=Skills,width=20,font="arial 10")
skills_entry.place(x=630,y=150)

Semester=Combobox(obj,values=['SEM 1','SEM 2','SEM 3','SEM 4','SEM 5','SEM 6'],font="Roboto 10",width=17,state='r')
Semester.place(x=630,y=50)
Semester.set('Select Semester')


#parent details
obj2=LabelFrame(root,text="Parent's Detail",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=220,relief=GROOVE)
obj2.place(x=30,y=470)

Label(obj2,text="Father Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)

Label(obj2,text="Mother Name:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)

Fname=StringVar()
fname_entry=Entry(obj2,textvariable=Fname,width=20,font="arial 10")
fname_entry.place(x=160,y=50)

F_occ_name=StringVar()
f_occ_name_entry=Entry(obj2,textvariable=F_occ_name,width=20,font="arial 10")
f_occ_name_entry.place(x=160,y=100)

Mname=StringVar()
mname_entry=Entry(obj2,textvariable=Mname,width=20,font="arial 10")
mname_entry.place(x=630,y=50)

M_occ_name=StringVar()
m_occ_name_entry=Entry(obj2,textvariable=M_occ_name,width=20,font="arial 10")
m_occ_name_entry.place(x=630,y=100)

#image
f=Frame(root,bd=3,bg='black',width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)

img=PhotoImage(file="upload.png")
lbl=Label(f,image=img)
lbl.place(x=0,y=0)

#button
Button(root,text="Upload",width=19,height=2,font="arial 12 bold",bg="lightblue",command=showimage).place(x=1000,y=370)

saveButton=Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="lightgreen",command=Save)
saveButton.place(x=1000,y=450)

Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="lightpink",command=Clear).place(x=1000,y=530)

Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="gray",command=Exit).place(x=1000,y=610)

root.mainloop()
